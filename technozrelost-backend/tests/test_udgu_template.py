"""Тесты шва udgu-template: шаблон, ТЗ и пример архива — публичный интерфейс.

Почему на шве: проверяем существование файлов, открываемость Excel,
листы/заголовки/примеры/валидацию и содержимое ZIP через файловую систему,
без БД и внутренностей генерации. Ожидаемые значения — известные величины
из ТЗ, не вычисленные кодом под тестом.
"""

from __future__ import annotations

import pathlib
import zipfile

import openpyxl


def _base() -> pathlib.Path:
    # docs/udgu_template относительно technozrelost-backend
    return pathlib.Path("docs/udgu_template")


def test_udgu_template_files_exist() -> None:
    """Файлы шва существуют: xlsx, pdf, README, schema, example ZIP."""
    base = _base()
    xlsx = base / "00_опись.xlsx"
    pdf = base / "00_опись.pdf"
    readme = base / "README_УдГУ_выгрузка.md"
    schema = base / "schema" / "udgu_template.schema.json"
    zip_path = base / "example" / "УдГУ_потенциалУР_2026-09-03.zip"

    for path in (xlsx, pdf, readme, schema, zip_path):
        assert path.exists(), f"отсутствует {path}"
        assert path.stat().st_size > 0, f"пустой {path}"

    assert pdf.read_bytes()[:4] == b"%PDF", "pdf должен начинаться с %PDF"


def test_udgu_template_xlsx_opens_and_has_sheets() -> None:
    """Excel открывается, 9+ листов с ожидаемыми именами."""
    xlsx = _base() / "00_опись.xlsx"
    wb = openpyxl.load_workbook(str(xlsx))
    expected = [
        "00_инструкция",
        "01_кафедры_лаб",
        "02_приоритеты_заделы",
        "03_миссия_фронтир",
        "04_оборудование",
        "05_РИД",
        "06_услуги_МСП",
        "07_люди_эксперты",
        "99_raw_опись",
    ]
    assert len(wb.sheetnames) >= 9, f"ожидается 9+ листов, получили {wb.sheetnames}"
    for name in expected:
        assert name in wb.sheetnames, f"отсутствует лист {name}"


def test_udgu_template_headers_example_and_validation() -> None:
    """Заголовки, пример, подсказки и валидация — известные величины."""
    xlsx = _base() / "00_опись.xlsx"
    wb = openpyxl.load_workbook(str(xlsx))

    header_expectations: dict[str, list[str]] = {
        "01_кафедры_лаб": [
            "название*",
            "факультет",
            "руководитель",
            "компетенции",
            "файл_приложение",
        ],
        "02_приоритеты_заделы": [
            "направление",
            "TRL",
            "компетенции",
            "файл_приложение",
        ],
        "03_миссия_фронтир": [
            "миссия",
            "фронтир",
            "стратегия",
            "файл_приложение",
        ],
        "04_оборудование": [
            "название*",
            "год_выпуска",
            "доступно_для_МСП",
            "компетенции",
            "файл_приложение",
        ],
        "05_РИД": [
            "название*",
            "тип",
            "дата",
            "авторы",
            "файл_приложение",
        ],
        "06_услуги_МСП": [
            "название*",
            "формат",
            "компетенции",
            "файл_приложение",
        ],
        "07_люди_эксперты": [
            "ФИО*",
            "должность",
            "компетенции",
            "файл_приложение",
        ],
        "99_raw_опись": ["файл*", "тип_сущности", "описание"],
    }

    for sheet, substrings in header_expectations.items():
        ws = wb[sheet]
        headers = [c.value for c in ws[1] if c.value]
        assert headers, f"пустые заголовки в {sheet}"
        for needle in substrings:
            assert any(
                needle.lower() in str(h).lower() for h in headers
            ), f"в {sheet} нет '{needle}' — {headers}"
        example = [c.value for c in ws[2]]
        assert any(v not in (None, "") for v in example), f"пустая пример-строка в {sheet}"
        text = " ".join(str(v) for v in example if v)
        assert len(text) > 10, f"короткий пример в {sheet}: {text}"
        comments = [c.comment for c in ws[1] if c.comment]
        assert len(comments) >= 2, f"ожидаются подсказки в {sheet}, нашли {len(comments)}"

    ws_priority = wb["02_приоритеты_заделы"]
    dvs = list(ws_priority.data_validations.dataValidation)
    assert len(dvs) >= 1, "в 02 ожидается валидация TRL 1-9"

    ws_equip = wb["04_оборудование"]
    dvs = list(ws_equip.data_validations.dataValidation)
    assert len(dvs) >= 2, "в 04 ожидается валидация года и Да/Нет"

    ws_rid = wb["05_РИД"]
    dvs = list(ws_rid.data_validations.dataValidation)
    assert len(dvs) >= 1, "в 05 ожидается валидация типа РИД"

    ws_mission = wb["03_миссия_фронтир"]
    dvs = list(ws_mission.data_validations.dataValidation)
    assert len(dvs) >= 1, "в 03 ожидается валидация стратегии года"

    ws_instr = wb["00_инструкция"]
    instr_text = " ".join(
        str(c.value) for row in ws_instr.iter_rows() for c in row if c.value
    )
    assert "нет данных" in instr_text.lower(), "инструкция: 'нет данных'"
    has_empty = "пустая" in instr_text.lower() or "пустой" in instr_text.lower()
    assert has_empty, "инструкция: пустая папка"


def test_udgu_template_readme_contains_required_sections() -> None:
    """README_УдГУ_выгрузка.md — известные разделы ТЗ, не из кода."""
    readme = _base() / "README_УдГУ_выгрузка.md"
    text = readme.read_text(encoding="utf-8")
    must = [
        "реестр потенциала УР",
        "ZIP",
        "01_кафедры",
        "04_оборудование",
        "05_РИД",
        "07_люди",
        "таблички",
        "ядро",
        "документы",
        "приложения",
        "08_",
        "Wave1",
        "Wave2",
        "УдГУ_потенциалУР_YYYY-MM-DD.zip",
        "маппинг",
    ]
    lower = text.lower()
    for needle in must:
        assert needle.lower() in lower, f"в README отсутствует '{needle}'"
    assert "Equipment" in text or "equipment" in text.lower()
    assert "competencies" in text.lower()


def test_udgu_template_example_zip_unpacks_and_has_folders() -> None:
    """Example ZIP распаковывается и содержит ожидаемые папки."""
    zip_path = _base() / "example" / "УдГУ_потенциалУР_2026-09-03.zip"
    assert zip_path.exists()
    with zipfile.ZipFile(str(zip_path), "r") as zf:
        assert zf.testzip() is None, "архив повреждён (CRC)"
        names = zf.namelist()
        must_prefixes = [
            "01_кафедры_и_лаборатории/",
            "02_приоритеты_и_заделы/",
            "03_миссия_фронтир/",
            "04_оборудование/",
            "05_РИД",
            "06_услуги_МСП/",
            "07_люди_эксперты/",
            "08_extra_example/",
            "raw/",
        ]
        for pref in must_prefixes:
            assert any(n.startswith(pref) for n in names), f"нет префикса {pref}"
        has_readme = any(
            n in ("README.md", "README_УдГУ_выгрузка.md") for n in names
        )
        assert has_readme, "в корне ZIP должен быть README"
        assert any(n == "00_опись.xlsx" for n in names), "нет 00_опись.xlsx"
        for pref in must_prefixes:
            files_in = [n for n in names if n.startswith(pref) and not n.endswith("/")]
            assert len(files_in) >= 1, f"папка {pref} пуста"
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            zf.extractall(tmp)
            extracted = list(pathlib.Path(tmp).rglob("*"))
            assert len(extracted) >= 10, "мало файлов после распаковки"
            xlsx_in_zip = pathlib.Path(tmp) / "00_опись.xlsx"
            assert xlsx_in_zip.exists()
            wb = openpyxl.load_workbook(str(xlsx_in_zip))
            assert len(wb.sheetnames) >= 9
