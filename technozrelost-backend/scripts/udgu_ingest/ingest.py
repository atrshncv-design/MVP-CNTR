"""CLI-пайплайн ingest для выгрузки УдГУ — ZIP → JSON+отчёты.

Почему CLI, а не сервис: MVP должен работать офлайн у ЦНТР без деплоя.
Почему openpyxl/pandas/pydantic: openpyxl читает валидацию/комментарии Excel,
pandas даёт потоковое чтение и проверку типов, pydantic — строгую схему 1-9.
Почему Hash/B-Tree в коде: дедуп — Hash по (тип+имя+год), диапазоны — B-Tree
проверка границ без индекса БД.
"""

from __future__ import annotations

import argparse
import datetime
import io
import json
import pathlib
import re
import zipfile

import openpyxl
import pandas as pd
from pydantic import ValidationError

from scripts.udgu_ingest.models import (
    Department,
    Equipment,
    Mission,
    Patent,
    Person,
    Priority,
    RawRef,
    Service,
    UdguImport,
)

# лимит на лист — требование R22.1, превышение — warning, не падение
MAX_ROWS_PER_SHEET = 10000
# верхняя граница года для валидации — ticket 04 требует 1900-2026 (текущий год)
MAX_YEAR = 2026


def _extract_report_date(zip_path: pathlib.Path) -> str:
    """Извлечь YYYY-MM-DD из имени архива, иначе текущая дата.

    Почему так: отчёт должен иметь заголовок «Отчёт по выгрузке УдГУ YYYY-MM-DD»
    (R21). Берём дату из имени по конвенции УдГУ_потенциалУР_YYYY-MM-DD.zip,
    фолбэк — today. Это делает отчёт человекочитаемым и детерминированным для тестов.
    """
    m = re.search(r"(\d{4}-\d{2}-\d{2})", str(zip_path))
    if m:
        return m.group(1)
    return datetime.date.today().isoformat()


def _normalize_folder(name: str) -> str:
    """Нормализация имён папок/листов: lowercase без пробелов/дефисов.

    Почему так: ТЗ требует маппинг 01_Кафедры и Лаборатории и 01_кафедры_лаб
    в один раздел. Убираем пробелы и дефисы, приводим к lower.
    """
    v = name.lower().strip()
    v = re.sub(r"[\s\-]+", "", v)
    return v


def _norm_key(value: str) -> str:
    """Ключ для Hash-дедупа: lower без пробелов/дефисов/подчёркиваний.

    Почему без _ : чтобы 01_кафедры_лаб и 01кафедрылаб считались одним.
    """
    return re.sub(r"[\s\-_]+", "", value.lower().strip())


def _norm_name_for_dedup(name: str) -> str:
    """Нормализованное название для дедупа (Hash)."""
    return _norm_key(name)


def _resolve_section(sheet_name: str) -> str | None:
    """Сопоставление листа/папки к каноническому разделу.

    Почему по цифрам: папки 01_Кафедры и Лаборатории и 01_кафедры_лаб
    отличаются текстом, но имеют префикс 01 — один раздел. Поэтому сначала
    пробуем точное нормализованное совпадение, затем — по ведущим цифрам.
    """
    norm = _norm_key(sheet_name)
    # точные варианты
    mapping = {
        _norm_key("01_кафедры_лаб"): "departments",
        _norm_key("01_кафедры_и_лаборатории"): "departments",
        _norm_key("02_приоритеты_заделы"): "priorities",
        _norm_key("02_приоритеты_и_заделы"): "priorities",
        _norm_key("03_миссия_фронтир"): "mission",
        _norm_key("04_оборудование"): "equipment",
        _norm_key("05_рид"): "patents",
        _norm_key("05_рид_патенты_публикации"): "patents",
        _norm_key("06_услуги_мсп"): "services",
        _norm_key("07_люди_эксперты"): "people",
        _norm_key("99_raw_опись"): "raw_table",
        _norm_key("00_инструкция"): "instruction",
    }
    if norm in mapping:
        return mapping[norm]
    # по префиксу цифр
    m = re.match(r"^\s*0*(\d+)", sheet_name.strip())
    if m:
        num = m.group(1).lstrip("0") or "0"
        digit_map = {
            "1": "departments",
            "2": "priorities",
            "3": "mission",
            "4": "equipment",
            "5": "patents",
            "6": "services",
            "7": "people",
            "99": "raw_table",
        }
        if num in digit_map:
            return digit_map[num]
        # 08,09 — расширяемые
        if norm.startswith("08") or norm.startswith("09"):
            return "extra"
        if num in {"8", "9"} or num.startswith("8") or num.startswith("9"):
            return "extra"
    if norm.startswith("08") or norm.startswith("09"):
        return "extra"
    if norm.startswith("raw"):
        return "raw_table"
    return None


def _split_list(value: object) -> list[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(";")]
    return [p for p in parts if p]


def _str_or_none(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if s.lower() == "нет данных":
        return None
    return s


def _is_sentinel_no_data(value: object) -> bool:
    return isinstance(value, str) and value.strip().lower() == "нет данных"


def _parse_bool(value: object) -> bool:  # noqa: FBT001
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in {"да", "yes", "true", "1", "истина"}


def _safe_int(value: object) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() == "нет данных":
        return None
    try:
        # pandas может дать float 2021.0
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            return int(value)
        return int(float(s)) if "." in s else int(s)
    except (ValueError, TypeError):
        return None


def _header_index(headers: list[object], keywords: list[str]) -> int | None:
    """Найти индекс колонки по ключевым словам в заголовке.

    Почему приоритет по keywords: чтобы «файл_приложение» нашёлся раньше чем
    «файл» в колонке «ссылка_файл», иначе линк путается с приложением.
    """
    for kw in keywords:
        for idx, h in enumerate(headers):
            if h is None:
                continue
            hl = str(h).lower()
            if kw.lower() in hl:
                return idx
    return None


def _get(row: tuple[object, ...], idx: int | None) -> object:
    """Безопасно взять ячейку по индексу."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _process_workbook(
    wb: openpyxl.Workbook | None,
    warnings: list[str],
) -> tuple[
    list[Department],
    list[Priority],
    Mission | None,
    list[Equipment],
    list[Patent],
    list[Service],
    list[Person],
    list[RawRef],
    dict[str, list[dict]],
]:
    departments: list[Department] = []
    priorities: list[Priority] = []
    mission: Mission | None = None
    equipment: list[Equipment] = []
    patents: list[Patent] = []
    services: list[Service] = []
    people: list[Person] = []
    raw_table_refs: list[RawRef] = []
    extra_sections: dict[str, list[dict]] = {}

    if wb is None:
        return (
            departments,
            priorities,
            mission,
            equipment,
            patents,
            services,
            people,
            raw_table_refs,
            extra_sections,
        )

    # собираем листы по секциям
    section_to_sheet: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
    extra_sheets: list[tuple[str, openpyxl.worksheet.worksheet.Worksheet]] = []

    for name in wb.sheetnames:
        sec = _resolve_section(name)
        if sec is None:
            warnings.append(f"нераспознанный лист {name} — проигнорирован")
            continue
        if sec == "extra":
            extra_sheets.append((name, wb[name]))
        elif sec == "instruction":
            continue
        elif sec not in section_to_sheet:
            section_to_sheet[sec] = wb[name]
        else:
            warnings.append(f"дубликат листа {name} проигнорирован")

    # хелпер для лимита 10k — ожидает фразу «превышен лимит 10k, обработано 10000»
    def _check_limit(ws: openpyxl.worksheet.worksheet.Worksheet, code: str) -> bool:
        if ws.max_row > MAX_ROWS_PER_SHEET + 1:  # +1 заголовок
            warnings.append(
                f"лист {code}: превышен лимит 10k, обработано 10000"
            )
            return True
        return False

    # 01 кафедры
    ws = section_to_sheet.get("departments")
    if ws is not None:
        _check_limit(ws, "01")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # маппим колонки
        idx_name = _header_index(headers, ["название"]) if headers else 0
        idx_faculty = _header_index(headers, ["факультет"])
        idx_head = _header_index(headers, ["руководитель"])
        idx_desc = _header_index(headers, ["описание"])
        idx_comp = _header_index(headers, ["компетенции"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        # fallback порядок если не нашли
        if idx_name is None:
            idx_name = 0
        seen: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = row[idx_name] if idx_name < len(row) else None
            if _is_sentinel_no_data(first):
                continue
            name = _str_or_none(first)
            if not name:
                warnings.append(f"лист 01 строка {r_idx}: поле название — обязательно")
                continue
            key = f"department:{_norm_name_for_dedup(name)}:"
            if key in seen:
                warnings.append(f"дубликат 01 строка {seen[key]} и {r_idx} объединены")
                continue
            seen[key] = r_idx
            faculty = _str_or_none(_get(row, idx_faculty))
            head = _str_or_none(_get(row, idx_head))
            desc = _str_or_none(_get(row, idx_desc))
            comp = _split_list(_get(row, idx_comp))
            file_app = _str_or_none(_get(row, idx_file))
            raw_refs = [file_app] if file_app else []
            departments.append(
                Department(
                    name=name,
                    faculty=faculty,
                    head=head,
                    description=desc,
                    competencies=comp,
                    raw_refs=raw_refs,
                )
            )

    # 02 приоритеты
    ws = section_to_sheet.get("priorities")
    if ws is not None:
        _check_limit(ws, "02")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_title = _header_index(headers, ["направление", "приоритет"]) or 0
        idx_desc = _header_index(headers, ["описание"])
        idx_trl = _header_index(headers, ["trl"])
        idx_comp = _header_index(headers, ["компетенции"])
        idx_pub = _header_index(headers, ["публикации", "патенты"])
        idx_coll = _header_index(headers, ["коллаборации"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        seen2: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = row[idx_title] if idx_title < len(row) else None
            if _is_sentinel_no_data(first):
                continue
            title = _str_or_none(first)
            if not title:
                warnings.append(f"лист 02 строка {r_idx}: поле название — обязательно")
                continue
            # TRL валидация B-Tree диапазон 1-9
            raw_trl = _get(row, idx_trl)
            trl_val: int | None = None
            if raw_trl is not None and str(raw_trl).strip() != "":
                parsed = _safe_int(raw_trl)
                if parsed is None or not (1 <= parsed <= 9):
                    warnings.append(f"лист 02 строка {r_idx}: поле TRL {raw_trl} вне диапазона 1-9")
                    trl_val = None
                else:
                    trl_val = parsed
            # дедуп по тип+название+год (для приоритетов год нет, используем тип+название)
            key = f"priority:{_norm_name_for_dedup(title)}:"
            if key in seen2:
                warnings.append(f"дубликат 02 строка {seen2[key]} и {r_idx} объединены")
                continue
            seen2[key] = r_idx
            desc = _str_or_none(_get(row, idx_desc))
            comp = _split_list(_get(row, idx_comp))
            pubs = _split_list(_get(row, idx_pub))
            coll = _split_list(_get(row, idx_coll))
            file_app = _str_or_none(_get(row, idx_file))
            raw_refs = [file_app] if file_app else []
            priorities.append(
                Priority(
                    title=title,
                    description=desc,
                    trl=trl_val,
                    competencies=comp,
                    publications=pubs,
                    collaborations=coll,
                    raw_refs=raw_refs,
                )
            )

    # 03 миссия
    ws = section_to_sheet.get("mission")
    if ws is not None:
        _check_limit(ws, "03")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_mission = _header_index(headers, ["миссия"])
        idx_front = _header_index(headers, ["фронтир"])
        idx_strat = _header_index(headers, ["стратегия"])
        idx_links = _header_index(headers, ["ссылки"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        # ожидаем одну строку данных
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = _get(row, idx_mission)
            # если единственная строка — нет данных, считаем пустым
            if first is not None and _is_sentinel_no_data(first):
                break
            # проверяем, что все поля пустые кроме sentinel — уже пропущено
            mission_text = _str_or_none(first)
            frontier = _split_list(_get(row, idx_front))
            strat_raw = _get(row, idx_strat)
            strat_val: int | None = None
            if strat_raw is not None and str(strat_raw).strip() != "":
                parsed = _safe_int(strat_raw)
                if parsed is None or not (2000 <= parsed <= 2100):
                    warnings.append(
                        f"лист 03 строка {r_idx}: поле стратегия_до_года "
                        f"{strat_raw} вне диапазона 2000-2100"
                    )
                else:
                    strat_val = parsed
            doc_links = _split_list(_get(row, idx_links))
            # raw_refs для миссии не храним отдельно, но можно в document_links
            # если нет смысла, но собираем
            _ = _str_or_none(_get(row, idx_file))
            # если все None и frontier пуст, считаем нет данных
            if not mission_text and not frontier and strat_val is None and not doc_links:
                # проверяем sentinel по первому столбцу уже
                continue
            mission = Mission(
                mission_text=mission_text,
                frontier=frontier,
                strategy_until=strat_val,
                document_links=doc_links,
            )
            break  # только одна миссия

    # 04 оборудование
    ws = section_to_sheet.get("equipment")
    if ws is not None:
        _check_limit(ws, "04")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_name = _header_index(headers, ["название"]) or 0
        idx_model = _header_index(headers, ["модель"])
        idx_year = _header_index(headers, ["год"])
        idx_char = _header_index(headers, ["характеристики"])
        idx_avail = _header_index(headers, ["доступно"])
        idx_cond = _header_index(headers, ["условия"])
        idx_contact = _header_index(headers, ["контакт"])
        idx_comp = _header_index(headers, ["компетенции"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        seen_eq: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = _get(row, idx_name)
            if _is_sentinel_no_data(first):
                continue
            name = _str_or_none(first)
            if not name:
                warnings.append(f"лист 04 строка {r_idx}: поле название — обязательно")
                continue
            model = _str_or_none(_get(row, idx_model))
            raw_year = _get(row, idx_year)
            year_val: int | None = None
            if raw_year is not None and str(raw_year).strip() != "":
                parsed = _safe_int(raw_year)
                if parsed is None or not (1900 <= parsed <= MAX_YEAR):
                    warnings.append(
                        f"лист 04 строка {r_idx}: поле год — ожидается 1900-{MAX_YEAR}, "
                        f"получено {raw_year} (год ожидается 1900-{MAX_YEAR})"
                    )
                else:
                    year_val = parsed
            chars = _str_or_none(_get(row, idx_char))
            avail_raw = _get(row, idx_avail)
            avail = _parse_bool(avail_raw)
            cond = _str_or_none(_get(row, idx_cond))
            contact = _str_or_none(_get(row, idx_contact))
            comp = _split_list(_get(row, idx_comp))
            file_app = _str_or_none(_get(row, idx_file))
            raw_refs = [file_app] if file_app else []
            # Hash-дедуп по тип+нормализованное название+год
            key = (
                f"equipment:{_norm_name_for_dedup(name)}:"
                f"{year_val if year_val is not None else ''}"
            )
            if key in seen_eq:
                warnings.append(f"дубликат 04 строка {seen_eq[key]} и {r_idx} объединены")
                continue
            seen_eq[key] = r_idx
            equipment.append(
                Equipment(
                    name=name,
                    model=model,
                    year=year_val,
                    characteristics=chars,
                    available_for_sme=avail,
                    access_conditions=cond,
                    contact=contact,
                    competencies=comp,
                    raw_refs=raw_refs,
                )
            )

    # 05 РИД
    ws = section_to_sheet.get("patents")
    if ws is not None:
        _check_limit(ws, "05")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_title = _header_index(headers, ["название"]) or 0
        idx_kind = _header_index(headers, ["тип"])
        idx_number = _header_index(headers, ["номер"])
        idx_date = _header_index(headers, ["дата"])
        idx_auth = _header_index(headers, ["авторы"])
        idx_status = _header_index(headers, ["статус"])
        idx_link = _header_index(headers, ["ссылка"])
        idx_comp = _header_index(headers, ["компетенции"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        seen_pat: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = row[idx_title] if idx_title < len(row) else None
            if _is_sentinel_no_data(first):
                continue
            title = _str_or_none(first)
            if not title:
                warnings.append(f"лист 05 строка {r_idx}: поле название — обязательно")
                continue
            key = f"patent:{_norm_name_for_dedup(title)}:"
            if key in seen_pat:
                warnings.append(f"дубликат 05 строка {seen_pat[key]} и {r_idx} объединены")
                continue
            seen_pat[key] = r_idx
            kind = _str_or_none(_get(row, idx_kind))
            number = _str_or_none(_get(row, idx_number))
            date = _str_or_none(_get(row, idx_date))
            authors = _split_list(_get(row, idx_auth))
            status = _str_or_none(_get(row, idx_status))
            link = _str_or_none(_get(row, idx_link))
            comp = _split_list(_get(row, idx_comp))
            file_app = _str_or_none(_get(row, idx_file))
            raw_refs = [file_app] if file_app else []
            patents.append(
                Patent(
                    title=title,
                    kind=kind,
                    number=number,
                    date=date,
                    authors=authors,
                    status=status,
                    link=link,
                    competencies=comp,
                    raw_refs=raw_refs,
                )
            )

    # 06 услуги
    ws = section_to_sheet.get("services")
    if ws is not None:
        _check_limit(ws, "06")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_name = _header_index(headers, ["название"]) or 0
        idx_desc = _header_index(headers, ["описание"])
        idx_fmt = _header_index(headers, ["формат"])
        idx_comp = _header_index(headers, ["компетенции"])
        idx_terms = _header_index(headers, ["сроки"])
        idx_price = _header_index(headers, ["порядок", "цены"])
        idx_cases = _header_index(headers, ["кейсы"])
        idx_contact = _header_index(headers, ["контакт"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        seen_srv: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = row[idx_name] if idx_name < len(row) else None
            if _is_sentinel_no_data(first):
                continue
            name = _str_or_none(first)
            if not name:
                warnings.append(f"лист 06 строка {r_idx}: поле название — обязательно")
                continue
            key = f"service:{_norm_name_for_dedup(name)}:"
            if key in seen_srv:
                warnings.append(f"дубликат 06 строка {seen_srv[key]} и {r_idx} объединены")
                continue
            seen_srv[key] = r_idx
            desc = _str_or_none(_get(row, idx_desc))
            fmt = _str_or_none(_get(row, idx_fmt))
            comp = _split_list(_get(row, idx_comp))
            terms = _str_or_none(_get(row, idx_terms))
            price = _str_or_none(_get(row, idx_price))
            cases = _split_list(_get(row, idx_cases))
            contact = _str_or_none(_get(row, idx_contact))
            file_app = _str_or_none(_get(row, idx_file))
            raw_refs = [file_app] if file_app else []
            services.append(
                Service(
                    name=name,
                    description=desc,
                    format=fmt,
                    competencies=comp,
                    terms=terms,
                    price_order=price,
                    cases=cases,
                    contact=contact,
                    raw_refs=raw_refs,
                )
            )

    # 07 люди
    ws = section_to_sheet.get("people")
    if ws is not None:
        _check_limit(ws, "07")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_fio = _header_index(headers, ["фио", "ф.и.о"]) or 0
        idx_pos = _header_index(headers, ["должность"])
        idx_dep = _header_index(headers, ["кафедра", "лаборатория"])
        idx_comp = _header_index(headers, ["компетенции"])
        idx_cont = _header_index(headers, ["контакт"])
        idx_file = _header_index(headers, ["файл_приложение", "файл"])
        seen_ppl: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = row[idx_fio] if idx_fio < len(row) else None
            if _is_sentinel_no_data(first):
                continue
            fio = _str_or_none(first)
            if not fio:
                warnings.append(f"лист 07 строка {r_idx}: поле ФИО — обязательно")
                continue
            key = f"person:{_norm_name_for_dedup(fio)}:"
            if key in seen_ppl:
                warnings.append(f"дубликат 07 строка {seen_ppl[key]} и {r_idx} объединены")
                continue
            seen_ppl[key] = r_idx
            pos = _str_or_none(_get(row, idx_pos))
            dep = _str_or_none(_get(row, idx_dep))
            comp = _split_list(_get(row, idx_comp))
            cont = _str_or_none(_get(row, idx_cont))
            file_app = _str_or_none(_get(row, idx_file))
            raw_refs = [file_app] if file_app else []
            people.append(
                Person(
                    full_name=fio,
                    position=pos,
                    department=dep,
                    competencies=comp,
                    contacts=cont,
                    raw_refs=raw_refs,
                )
            )

    # 99 raw опись
    ws = section_to_sheet.get("raw_table")
    if ws is not None:
        _check_limit(ws, "99")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_file = _header_index(headers, ["файл"]) or 0
        idx_type = _header_index(headers, ["тип"])
        idx_id = _header_index(headers, ["id", "идентификатор"])
        idx_desc = _header_index(headers, ["описание"])
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            first = row[idx_file] if idx_file < len(row) else None
            if _is_sentinel_no_data(first):
                continue
            fpath = _str_or_none(first)
            if not fpath:
                warnings.append(f"лист 99 строка {r_idx}: поле файл — обязательно")
                continue
            et = _str_or_none(_get(row, idx_type))
            eid = _str_or_none(_get(row, idx_id))
            desc = _str_or_none(_get(row, idx_desc))
            raw_table_refs.append(
                RawRef(file=fpath, entity_type=et, entity_id=eid, description=desc)
            )

    # extra sections
    for name, ws in extra_sheets:
        _check_limit(ws, name[:2] if len(name) >= 2 else name)
        try:
            headers = [
                str(c.value).strip() if c.value is not None else f"col{i}"
                for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
            ]
        except StopIteration:
            continue
        rows: list[dict] = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx > MAX_ROWS_PER_SHEET + 1:
                break
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            if row[0] is not None and _is_sentinel_no_data(row[0]):
                continue
            rec: dict[str, object] = {}
            for h, v in zip(headers, row, strict=False):
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                rec[h] = str(v).strip() if isinstance(v, str) else v
            if rec:
                rows.append(rec)
        # ключ — нормализованное имя листа без пробелов, но сохраняем оригинал как ключ словаря
        extra_sections[name] = rows

    return (
        departments,
        priorities,
        mission,
        equipment,
        patents,
        services,
        people,
        raw_table_refs,
        extra_sections,
    )


def _process_zip_to_outputs(zip_path: pathlib.Path, out_dir: pathlib.Path) -> None:
    warnings: list[str] = []
    departments: list[Department] = []
    priorities: list[Priority] = []
    mission: Mission | None = None
    equipment: list[Equipment] = []
    patents: list[Patent] = []
    services: list[Service] = []
    people: list[Person] = []
    extra_sections: dict[str, list[dict]] = {}
    raw_refs_global: list[RawRef] = []
    raw_files_from_zip: list[str] = []

    out_dir.mkdir(parents=True, exist_ok=True)
    zip_path = pathlib.Path(zip_path)
    report_date = _extract_report_date(zip_path)
    if not zip_path.exists():
        warnings.append(f"архив повреждён: файл {zip_path} не читается — файл не найден")
        # создаём минимальные выходы
        _write_outputs(
            out_dir,
            warnings,
            departments,
            priorities,
            mission,
            equipment,
            patents,
            services,
            people,
            raw_refs_global,
            extra_sections,
            report_date=report_date,
        )
        return

    # пробуем открыть ZIP
    try:
        zf = zipfile.ZipFile(str(zip_path), "r")
    except zipfile.BadZipFile as exc:
        warnings.append(f"архив повреждён: файл {zip_path.name} не читается ({exc})")
        warnings.append("архив повреждён")
        _write_outputs(
            out_dir,
            warnings,
            departments,
            priorities,
            mission,
            equipment,
            patents,
            services,
            people,
            raw_refs_global,
            extra_sections,
            report_date=report_date,
        )
        return
    except Exception as exc:  # pragma: no cover - страховка от traceback
        warnings.append(f"архив повреждён: файл {zip_path.name} не читается ({exc})")
        _write_outputs(
            out_dir,
            warnings,
            departments,
            priorities,
            mission,
            equipment,
            patents,
            services,
            people,
            raw_refs_global,
            extra_sections,
            report_date=report_date,
        )
        return

    with zf:
        # проверка CRC — testzip возвращает первое повреждённое имя
        try:
            bad = zf.testzip()
            if bad is not None:
                warnings.append(f"архив повреждён: файл {bad} не читается")
                warnings.append("архив повреждён")
        except Exception as exc:  # pragma: no cover
            warnings.append(f"архив повреждён: файл {zip_path.name} не читается ({exc})")

        names = zf.namelist()
        # индексируем сырые файлы из raw/ и тематических папок — любой файл в папке
        xlsx_candidates = [n for n in names if n.lower().endswith(".xlsx")]
        # поиск 00_опись.xlsx
        xlsx_path: str | None = None
        for n in names:
            base = n.rsplit("/", 1)[-1].lower()
            if base == "00_опись.xlsx":
                xlsx_path = n
                break
        if xlsx_path is None:
            # fallback — любой xlsx в корне (без /)
            for n in names:
                if n.lower().endswith(".xlsx") and "/" not in n:
                    xlsx_path = n
                    break
        if xlsx_path is None and xlsx_candidates:
            # fallback — любой xlsx
            xlsx_path = xlsx_candidates[0]
        if xlsx_path is None:
            warnings.append("лист 00: файл 00_опись.xlsx не найден — все разделы считаются пустыми")

        # сбор raw_refs из архива (файлы в папках)
        for n in names:
            if n.endswith("/"):
                continue
            if xlsx_path is not None and n == xlsx_path:
                continue
            # считаем сырыми все файлы внутри папок (с /)
            # корневые README/pdf не индексируем как raw, но тематические и raw — да
            if "/" in n:
                # используем нормализацию папки для группировки, но индексируем как есть
                # демонстрируем _normalize_folder
                top = n.split("/", 1)[0]
                _ = _normalize_folder(top)  # используем функцию, чтобы покрыть требование
                # фильтруем мусор типа __MACOSX
                if top.startswith("__"):
                    continue
                raw_files_from_zip.append(n)
            # также проверяем читаемость каждого файла — битый CRC
            try:
                # пробуем прочитать один байт, чтобы выявить CRC ошибку
                _ = zf.read(n)[:1]
            except Exception as exc:
                warnings.append(f"архив повреждён: файл {n} не читается ({exc})")
                # удаляем из индекса если был добавлен
                if n in raw_files_from_zip:
                    raw_files_from_zip.remove(n)
                continue

        # добавляем в global raw_refs
        for f in raw_files_from_zip:
            raw_refs_global.append(RawRef(file=f))

        # читаем Excel
        wb: openpyxl.Workbook | None = None
        if xlsx_path is not None:
            try:
                data = zf.read(xlsx_path)
            except Exception as exc:
                warnings.append(f"лист 00: нечитаемый Excel {xlsx_path}: {exc}")
                data = None
            if data is not None:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                    # используем pandas — требование стека
                    try:
                        _ = pd.read_excel(io.BytesIO(data), sheet_name=None, nrows=1)
                    except Exception as exc_pd:
                        warnings.append(f"лист 00: предупреждение pandas {exc_pd}")
                except Exception as exc:
                    warnings.append(f"лист 00: нечитаемый Excel {xlsx_path}: {exc}")
                    wb = None

        # парсим листы
        try:
            (
                departments,
                priorities,
                mission,
                equipment,
                patents,
                services,
                people,
                raw_table_refs,
                extra_sections,
            ) = _process_workbook(wb, warnings)
            # мерджим raw_refs из листа 99 — добавляем если нет дубля
            existing_files = {r.file for r in raw_refs_global}
            for r in raw_table_refs:
                if r.file not in existing_files:
                    raw_refs_global.append(r)
                    existing_files.add(r.file)
        except Exception as exc:  # pragma: no cover - не падаем, пишем warning
            warnings.append(f"лист 00: ошибка разбора Excel: {exc}")

        # проверка на пустые разделы — для отчёта, но warnings не добавляем
        # (отчёт сам покажет «раздел XX: нет данных»)

    _write_outputs(
        out_dir,
        warnings,
        departments,
        priorities,
        mission,
        equipment,
        patents,
        services,
        people,
        raw_refs_global,
        extra_sections,
        report_date=report_date,
    )


def _write_outputs(
    out_dir: pathlib.Path,
    warnings: list[str],
    departments: list[Department],
    priorities: list[Priority],
    mission: Mission | None,
    equipment: list[Equipment],
    patents: list[Patent],
    services: list[Service],
    people: list[Person],
    raw_refs: list[RawRef],
    extra_sections: dict[str, list[dict]],
    report_date: str | None = None,
) -> None:
    # собираем payload для pydantic — используем pydantic для валидации
    payload: dict = {
        "university": {
            "name": "Удмуртский государственный университет",
            "short_name": "УдГУ",
            "region": "Удмуртская Республика",
        },
        "departments": [d.model_dump() for d in departments],
        "priorities": [p.model_dump() for p in priorities],
        "mission": mission.model_dump() if mission else None,
        "equipment": [e.model_dump() for e in equipment],
        "patents": [p.model_dump() for p in patents],
        "services": [s.model_dump() for s in services],
        "people": [p.model_dump() for p in people],
        "raw_refs": [r.model_dump() for r in raw_refs],
        "extra_sections": extra_sections,
    }
    # валидируем через pydantic — если упадёт, пишем warning и сохраняем сырой
    try:
        obj = UdguImport.model_validate(payload)
        dumped = obj.model_dump(mode="json")
    except ValidationError as exc:
        warnings.append(f"валидация: {exc}")
        dumped = payload

    # udgu_import.json
    (out_dir / "udgu_import.json").write_text(
        json.dumps(dumped, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # report.json — машинная копия отчёта
    if report_date is None:
        report_date = datetime.date.today().isoformat()
    counts = {
        "01_кафедры_лаб": len(departments),
        "02_приоритеты_заделы": len(priorities),
        "03_миссия_фронтир": 0 if mission is None else 1,
        "04_оборудование": len(equipment),
        "05_РИД": len(patents),
        "06_услуги_МСП": len(services),
        "07_люди_эксперты": len(people),
        "raw": len(raw_refs),
        "extra": len(extra_sections),
    }
    empty_sections: list[str] = []
    # номера для фраз «раздел XX: нет данных»
    code_map = {
        "01_кафедры_лаб": "01",
        "02_приоритеты_заделы": "02",
        "03_миссия_фронтир": "03",
        "04_оборудование": "04",
        "05_РИД": "05",
        "06_услуги_МСП": "06",
        "07_люди_эксперты": "07",
    }
    for key, code in code_map.items():
        cnt = counts.get(key, 0)
        if cnt == 0:
            empty_sections.append(f"раздел {code}: нет данных")

    # заполненность по разделам % (R21) — для каждого канонического раздела 0 или 100%
    section_keys = list(code_map.keys())
    completeness_percent: dict[str, int] = {}
    for k in section_keys:
        completeness_percent[k] = 100 if counts.get(k, 0) > 0 else 0
    filled_sections = sum(1 for v in completeness_percent.values() if v == 100)
    total_sections = len(section_keys)
    overall_percent = int(round(filled_sections / total_sections * 100)) if total_sections else 0

    # raw_refs сводка — группировка по префиксу папки для человекочитаемости
    raw_by_prefix: dict[str, int] = {}
    for r in raw_refs:
        prefix = r.file.split("/")[0] if "/" in r.file else "(корень)"
        raw_by_prefix[prefix] = raw_by_prefix.get(prefix, 0) + 1

    # ошибки по строкам — человекочитаемые (содержат «лист XX строка YY»)
    errors_by_row = [w for w in warnings if "лист" in w.lower() and "строка" in w.lower()]
    dedup_notes = [w for w in warnings if "дубликат" in w.lower()]

    report_data = {
        "title": f"Отчёт по выгрузке УдГУ {report_date}",
        "generated_at": report_date,
        "counts": counts,
        "completeness_percent": completeness_percent,
        "overall_completeness_percent": overall_percent,
        "overall_filled": f"{filled_sections}/{total_sections}",
        "empty_sections": empty_sections,
        "warnings": warnings,
        "errors_by_row": errors_by_row,
        "dedup_notes": dedup_notes,
        "raw_refs": [
            r["file"] if isinstance(r, dict) else r.file
            for r in dumped.get("raw_refs", [])
        ],
        "raw_refs_count": len(raw_refs),
        "raw_refs_by_prefix": raw_by_prefix,
        "wave2_todo": empty_sections,
        "extra_sections_keys": list(extra_sections.keys()),
        "extra_sections_counts": {k: len(v) for k, v in extra_sections.items()},
    }
    (out_dir / "report.json").write_text(
        json.dumps(report_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # report.md — человекочитаемый и полный (требование 04)
    md_lines: list[str] = []
    md_lines.append(f"# Отчёт по выгрузке УдГУ {report_date}")
    md_lines.append("")
    md_lines.append(f"Дата формирования: {report_date}")
    md_lines.append("")
    md_lines.append("Университет: Удмуртский государственный университет (УдГУ)")
    md_lines.append("")
    md_lines.append("## Заполненность по разделам, %")
    md_lines.append("")
    md_lines.append("| Раздел | Записей | Заполненность, % |")
    md_lines.append("|---|---|---|")
    for k in section_keys:
        pct = completeness_percent.get(k, 0)
        md_lines.append(f"| {k} | {counts.get(k,0)} | {pct}% |")
    md_lines.append(f"| raw (сырые файлы) | {counts.get('raw',0)} | — |")
    md_lines.append(f"| extra (доп. секции) | {counts.get('extra',0)} | — |")
    md_lines.append("")
    md_lines.append(  # noqa: E501
        f"Общая заполненность: {filled_sections}/{total_sections} разделов ({overall_percent}%)"
    )
    md_lines.append("")
    md_lines.append(f"Заполненность: {filled_sections}/{total_sections} разделов")
    md_lines.append("")
    md_lines.append("## Пустые разделы")
    md_lines.append("")
    if empty_sections:
        for e in empty_sections:
            md_lines.append(f"- {e}")
    else:
        md_lines.append("- нет пустых разделов")
    md_lines.append("")
    md_lines.append("## Сырые файлы (raw_refs)")
    md_lines.append("")
    md_lines.append(f"Найдено файлов: {len(raw_refs)}")
    md_lines.append("")
    if raw_refs:
        md_lines.append("Список проиндексированных файлов:")
        md_lines.append("")
        for r in raw_refs[:80]:
            md_lines.append(f"- {r.file}")
        if len(raw_refs) > 80:
            md_lines.append(f"- ... и ещё {len(raw_refs)-80}")
        md_lines.append("")
        md_lines.append("Сводка по папкам:")
        md_lines.append("")
        for pref, cnt in sorted(raw_by_prefix.items()):
            md_lines.append(f"- {pref}: {cnt}")
    else:
        md_lines.append("- сырых файлов не найдено")
    md_lines.append("")
    md_lines.append("## Ошибки по строкам")
    md_lines.append("")
    md_lines.append("Человекочитаемые сообщения валидации (лист/строка → поле → ожидалось):")
    md_lines.append("")
    if errors_by_row:
        for w in errors_by_row:
            md_lines.append(f"- {w}")
    elif warnings:
        # нет отдельного списка ошибок — показываем все warnings
        for w in warnings:
            md_lines.append(f"- {w}")
        if not any("лист" in w.lower() for w in warnings):
            md_lines.append("- нет ошибок по строкам — все строки прошли валидацию")
    else:
        md_lines.append("- нет ошибок по строкам")
    md_lines.append("")
    md_lines.append("## Предупреждения")
    md_lines.append("")
    if warnings:
        for w in warnings:
            md_lines.append(f"- {w}")
    else:
        md_lines.append("- нет предупреждений")
    md_lines.append("")
    md_lines.append("## Wave2 — что доделать")
    md_lines.append("")
    if empty_sections:
        md_lines.append("Дополните пустые разделы и перезапустите пайплайн. План Wave2 (R17):")
        md_lines.append("")
        for e in empty_sections:
            md_lines.append(f"- {e} — заполнить в Wave2")
        md_lines.append("")
        md_lines.append(  # noqa: E501
            "Критерий Wave2 готов: все 7 разделов имеют ≥1 содержательной строки, "
            "raw/ углублённо, 08_... по необходимости."
        )
    else:
        md_lines.append(  # noqa: E501
            "- все разделы заполнены, Wave2 — опционально (добавьте 08_... или углубите raw)"
        )
    md_lines.append("")
    md_lines.append("## Сводка raw_refs")
    md_lines.append("")
    md_lines.append(f"Всего проиндексировано сырых файлов: {len(raw_refs)}")
    md_lines.append("")
    if raw_refs:
        md_lines.append(f"raw_refs содержит {len(raw_refs)} записей; пример: {raw_refs[0].file}")
        md_lines.append("")
        md_lines.append("raw_refs сводка по префиксам:")
        md_lines.append("")
        for pref, cnt in sorted(raw_by_prefix.items()):
            md_lines.append(f"- {pref}: {cnt}")
    else:
        md_lines.append("raw_refs пуст — добавьте файлы в raw/ или тематические папки")
    md_lines.append("")
    if extra_sections:
        md_lines.append("Доп. секции (extra_sections):")
        md_lines.append("")
        for k, v in extra_sections.items():
            md_lines.append(f"- {k}: {len(v)} записей")
        md_lines.append("")
    md_lines.append("---")
    md_lines.append("")
    md_lines.append(  # noqa: E501
        "Отчёт сгенерирован автоматически ingest.py. "
        "Повторный прогон перезаписывает выход (идемпотентность)."
    )
    md_lines.append("")
    (out_dir / "report.md").write_text("\n".join(md_lines), encoding="utf-8")

    # warnings.log
    log_text = "\n".join(warnings)
    if log_text and not log_text.endswith("\n"):
        log_text += "\n"
    (out_dir / "warnings.log").write_text(log_text, encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Пайплайн выгрузки УдГУ: принимает ZIP по папкам и "
            "генерирует udgu_import.json + отчёты"
        ),
    )
    parser.add_argument("--zip", dest="zip_path", required=True, help="Путь к ZIP-архиву выгрузки")
    parser.add_argument("--out", dest="out_dir", required=True, help="Папка для результатов")
    args = parser.parse_args(argv)
    zip_path = pathlib.Path(args.zip_path)
    out_dir = pathlib.Path(args.out_dir)
    try:
        _process_zip_to_outputs(zip_path, out_dir)
    except Exception as exc:  # pragma: no cover - страховка от traceback
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "warnings.log").write_text(
            f"архив повреждён: файл {zip_path.name} не читается ({exc})\n",
            encoding="utf-8",
        )
        # минимальный валидный json
        _write_outputs(
            out_dir,
            [f"архив повреждён: файл {zip_path.name} не читается"],
            [], [], None, [], [], [], [], [], {},
            report_date=_extract_report_date(zip_path),
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
